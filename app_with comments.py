import sys
import win32com.client as w3c
import os
import time
from datetime import date
import openpyxl
import pandas as pd
import datetime
import re

filename = os.path.dirname(os.path.abspath(__file__)) + "\workbook.xlsx"

#openpyxl easier to manipulate excel, but cannot unlock pw, so i unlock first with win32com then use openpyxl to modify, then use win32com to save with pw

#checks if employee can access info, id and pw from welcome function
def check(id, pw, filename):
    xlApp = w3c.Dispatch("Excel.Application") #starts Excel in the background
    try:
        wb = xlApp.Workbooks.Open(filename, False, False, None, pw) #opens file using the password entered in welcome function
    except:
        return "Wrong password or file path"
    xlApp.DisplayAlerts = False #suppresses Excel messages, choose default response to prompts
    ws = wb.Sheets(1) # counts from 1, not from 0
    try:
        xlUp = -4162 #vba command to move in up direction
        lastrow = ws.Cells(ws.Rows.Count, "A").End(xlUp).Row + 1 #find very last row in "A" col, move up till last row with content, +1 to find first empty cell
        
        for i in range(2,lastrow): # check if person is HR or not
            # LOOP RANGE OBJ
            if (ws.Range("A" + str(i)).Value) == id:
                if (ws.Range("D" + str(i)).Value) == "HR":
                    wb.SaveAs(filename, None, "", "") # if emp is HR, save file without password
                    xlApp.Quit()
                    return "isHR"
                else:
                    wb.SaveAs(filename, None, "hackerman", "")
                    xlApp.Quit()
                    return "notHR"

    except Exception as e:
        return "error"
    
    wb.SaveAs(filename, None, "hackerman", "") #if not found means is not in system
    xlApp.Quit()
    return "notInSys"

# welcome and login
def welcome():
    txt = "Welcome to the HR Birthday Program"
    print(txt.center(80, "-"))
    print()
    time.sleep(0.5) #prints welcome message before data and time
    df = None
    
    today = date.today()
    print("Today's date:", today.strftime("%B %d, %Y"))
    
    t = time.localtime()
    current_time = time.strftime("%H:%M:%S", t)
    print("Current time:", current_time)
    print()
    print("-"*80)
    print()
    
    while True:
        try:
            empid = int(input("Please input your employee ID: "))
        except:
            print("Invalid id, please try again")
            continue
        break
    
    print()
    print("Please input the password to unprotect the excel workbook, you have 3 tries")
    password = input("Try 1: ")

    try: #checking if person can access
        result = check(empid, password, filename)
    except Exception as e:
        print(e)
    
    for i in range(2):
        if (result == "Wrong password or file path"):
                print("Wrong password or file path, please try again")
                password = input(f"Try {i+2}: ")
                result = check(empid, password, filename)
        else:
            break
    if (result == "Wrong password or file path"):
        print("Out of tries, try again next time")
        print()
        print("-"*80)
        return 0, password #note the return values
    
    print()
    
    if (result == "isHR"):
        print("You are granted access")
        print()
        #get to home screen
        return 1, password #note the return values

    if (result != "isHR"):
        if (result == "notHR"):
            print("Sorry, access only for HR personnel.")
            print("Password has been changed, please contact IT personnel for assistance")
        elif (result == "error"):
            print("An error occurred")
        elif (result == "notInSys"):
            print("Sorry, access only for employees")
            print("Password has been changed, please contact IT personnel for assistance")
    
    print("-"*80)
    return 0, password #note the return values

# gets birthdays of people in specified groups
def get_bdays():
    df = pd.read_excel(filename) # for reading file, i just use pandas
    months = ["1","2","3","4","5","6","7","8","9","10","11","12","All"]
    depts = ["Finance", "HR", "IT", "All"]
    print("Which birthday month would you like to view?")
    print("Select from: Jan(1), Feb(2), Mar(3), Apr(4), ..., Dec(12), All")
    time.sleep(0.5)
    month = input("Please select month: ")
    print()
    while month not in months:
        print("Invalid month, please try again")
        print("Which birthday month would you like to view?")
        print("Select from: Jan(1), Feb(2), Mar(3), Apr(4), ..., Dec(12), All")
        time.sleep(0.5)
        month = input("Please select month: ")
        print()
    print("Which department would you like to view?")
    print("Select from: Finance, HR, IT, All")
    time.sleep(0.5)
    dept = input("Please select department: ")
    print()
    while dept not in depts:
        print("Invalid department, please try again")
        print("Which department would you like to view?")
        time.sleep(0.5)
        print("Select from: Finance, HR, IT, All")
        dept = input("Please select department: ")
        print()
    if (month == "All") and (dept == "All"): #if want to view everyone
        df2 = df.sort_values(["Birthday", "Last Name"], ascending = [True, True])
    elif (month == "All"): #view all months but specific dept
        df2 = df[df["Department"] == dept].sort_values(["Birthday", "Last Name"], ascending = [True, True])
    elif (dept == "All"): #view all depts but specific month
        df2 = df[(pd.to_datetime(df["Birthday"]).dt.month == int(month))].sort_values(["Birthday", "Last Name"], ascending = [True, True])
    else: # view specific month and dept
        df2 = df[(pd.to_datetime(df["Birthday"]).dt.month == int(month)) & (df["Department"] == dept)].sort_values(["Birthday", "Last Name"], ascending = [True, True])
    return df2.reset_index(drop = True)

# print birthdays from dataframe (dataframe from get_bdays())
def print_bdays(df):
    rows = df.shape[0]
    df["Birthday"]=df["Birthday"].astype(str) #convert to str for printing
    print(f"{rows} people have birthdays in this period")
    print()
    print('{:30s} {:30s} {:30s} '.format("First Name","Last Name","Birthday"))
    print(90*"-")
    for i in range(rows):
        row = df.iloc[i]
        print('{:30s} {:30s} {:30s}'.format(row["First Name"], row["Last Name"], row["Birthday"]))

# add employee record
def add_record():
    wb = openpyxl.load_workbook(filename) #using openpyxl now, different from win32com
    ws = wb.worksheets[0]
    depts = ["Finance", "HR", "IT"]
    celebrates = ["Y","N"]
    maxrow = ws.max_row
    repeat = 0
    while True: #check if employee id already exists
            try:
                empid = int(input("Please input employee ID to be added (enter -1 to quit): "))
                if empid == -1:
                    return
                for i in range(maxrow):
                    if empid == ws["A" + str(i+1)].value:
                        print("Employee id already exists, please try again")
                        repeat = 1
                        break
                if repeat == 1:
                    repeat = 0
                    continue
            except:
                print("Invalid id, please try again")
                continue

            break
    newrow = ws.max_row + 1 #find new row number for adding emp info
    ws["A" + str(newrow)].value = empid
    firstname = input("Please input employee's first name (enter -1 to quit): ")
    if firstname == "-1":
        return
    ws["B" + str(newrow)].value = firstname
    lastname = input("Please input employee's last name (enter -1 to quit): ")
    if lastname == "-1":
        return
    ws["C" + str(newrow)].value = lastname
    while True:
        dept = input("Please input employee's department (Finance, HR or IT) (enter -1 to quit): ")
        if dept == "-1":
            return
        if dept not in depts:
            print("Invalid dept, please try again")
            continue
        break
    ws["D" + str(newrow)].value = dept
    while True:
            try:
                bday = input("Please input employee's birthday (YYYY-MM-DD) (enter -1 to quit): ")
                if bday == "-1":
                    return
                empbday = datetime.datetime.strptime(bday, '%Y-%m-%d') #check valid date, if invalid will have error
            except:
                print("Invalid date, please try again")
                continue
            break
    ws["E" + str(newrow)].value = empbday.date()
    while True:
            try:
                datejoined = input("Please input employee's date joined (YYYY-MM-DD) (enter -1 to quit): ")
                if datejoined == "-1":
                    return
                empdatejoined = datetime.datetime.strptime(datejoined, '%Y-%m-%d')
            except:
                print("Invalid date, please try again")
                continue
            break
    ws["F" + str(newrow)].value = empdatejoined.date()
    while True:
        celebrate = input("Does employee wish to celebrate his/her birthday (Y or N) (enter -1 to quit): ")
        if celebrate == "-1":
            return
        if celebrate not in celebrates:
            print("Invalid value, please try again")
            continue
        break
    ws["G" + str(newrow)].value = celebrate
    wb.save(filename) #save file
    wb.close() #close file

# remove emp record
def remove_record():
    wb = openpyxl.load_workbook(filename)
    ws = wb.worksheets[0]
    removes = ["Y","N"]
    remove = "Y"
    while (remove != "N"):
        maxrow = ws.max_row # update max row each time looking for new emp to remove
        emprow = -1 # default emp row, will change if emp found
        repeat = 1 # default repeat value
        while True:
                    try:
                        empid = int(input("Please input employee ID you wish to remove (enter -1 to quit): "))
                        if empid == -1:
                            return
                        for i in range(maxrow): 
                            if empid == ws["A" + str(i+1)].value: # empid found
                                repeat = 0
                                emprow = i+1 # emp is at row i+1 (cos excel is i+1)
                                break
                        if repeat == 1:
                            print("Employee not found, please try again")
                            continue #loop again if not found
                    except:
                        print("Invalid id, please try again")
                        continue

                    break
        while True:
                remove = input("Employee found, are you sure you wish to delete the record (Y or N) (enter -1 to quit): ")
                if remove == "-1":
                    return
                if remove not in removes:
                    print("Invalid value, please try again")
                    continue
                break
        if (remove == "Y"):
            ws.delete_rows(emprow) #delete emp record
            wb.save(filename) #save file
            print("Employee record removed!")
            while True:
                remove = input("Remove another employee record (Y or N): ")
                if remove not in removes:
                    print("Invalid value, please try again")
                    continue
                break
        elif (remove == "N"):
            wb.close() # if not deleting any more, close file
            return
    wb.close() # close file
    return

# edit emp record
def edit_record():
    wb = openpyxl.load_workbook(filename)
    ws = wb.worksheets[0]
    yesorno = ["Y","N"]
    fields = ["FN", "LN", "D", "B", "DJ", "C"]
    depts = ["Finance", "HR", "IT"]
    edits = ["HS", "DE"]
    edit = "DE"
    #regular exp to check if value is indeed an actual date
    reg = r"(^(((0[1-9]|1[0-9]|2[0-8])[\/](0[1-9]|1[012]))|((29|30|31)[\/](0[13578]|1[02]))|((29|30)[\/](0[4,6,9]|11)))[\/](19|[2-9][0-9])\d\d$)|(^29[\/]02[\/](19|[2-9][0-9])(00|04|08|12|16|20|24|28|32|36|40|44|48|52|56|60|64|68|72|76|80|84|88|92|96)$)"
    while (edit != "HS"): # will look for new emp if edit != "HS"
        maxrow = ws.max_row # update maxrow each time u look for empid
        emprow = -1
        repeat = 1
        curedit = "Y"
        while True:
                try: # find empid
                    empid = int(input("Please input employee ID of record you wish to edit/view (enter -1 to quit): "))
                    if empid == -1:
                        return
                    for i in range(maxrow):
                        if empid == ws["A" + str(i+1)].value:
                            repeat = 0
                            emprow = i+1
                            break
                    if repeat == 1:
                        print("Employee not found, please try again")
                        continue
                except:
                    print("Invalid id, please try again")
                    continue

                break
        
        while (curedit == "Y"):
                print()
                print('{:20s} {:20s} {:20s} {:20s} {:20s} {:20s}'.format("First Name","Last Name","Department","Birthday","Date Joined","Celebrate?"))
                print(120*"-")
                # get all the current values in emp row
                fn = ws["B" + str(emprow)].value
                ln = ws["C" + str(emprow)].value
                d = ws["D" + str(emprow)].value
                dt = pd.to_datetime(ws["E" + str(emprow)].value)
                bd = dt.strftime("%d-%m-%Y")
                dt = pd.to_datetime(ws["F" + str(emprow)].value)
                dj = dt.strftime("%d-%m-%Y")
                c = ws["G" + str(emprow)].value
                print('{:20s} {:20s} {:20s} {:20s} {:20s} {:20s}'.format(fn, ln, d, bd, dj, c))
                print()
                print("Employee found, which field would you like to edit?")
                print("Select from: First Name(FN), Last Name(LN), Department(D), Birthday(B), Date Joined(DJ), Celebrate?(C)")
                print()
                while True:
                    field = input("Please select field (enter -1 to quit): ")
                    if field == "-1":
                        return
                    if field not in fields:
                        print("Invalid value, please try again")
                        continue
                    break
                print()
                while True:
                    if (field == "FN" or field == "LN"):
                        newval = input("Please enter new name: ")
                    elif (field == "D"):
                        newval = input("Please enter new department (Finance, HR or IT): ")
                        if newval not in depts:
                            print("Invalid department, please try again")
                            continue
                    elif (field == "B" or field == "DJ"):
                        newval = input("Please enter new date (DD/MM/YYYY): ")
                        if not (re.match(reg, newval)): # check if date provided is an actual date
                            print("Invalid date, please try again")
                            continue
                    elif (field == "C"):
                        newval = input("Please enter new celebration (Y or N): ")
                        if newval not in yesorno:
                            print("Invalid value, please try again")
                            continue
                    break
                print()
                while True:
                    apply = input("Apply changes (Y or N): ")
                    if (apply not in yesorno):
                        print("Invalid value, please try again")
                        continue
                    break
                if (apply == "N"):
                    while True:
                        curedit = input("Do you wish to edit further for the current employee (Y or N): ")
                        if (curedit not in yesorno):
                            print("Invalid value, please try again")
                            continue
                        break
                    continue
                if (apply == "Y"): # update values
                    if (field == "FN"):
                        ws["B" + str(emprow)].value = newval
                    elif (field == "LN"):
                        ws["C" + str(emprow)].value = newval
                    elif (field == "D"):
                        ws["D" + str(emprow)].value = newval
                    elif (field == "B"):
                        ws["E" + str(emprow)].value = newval
                    elif (field == "DJ"):
                        ws["F" + str(emprow)].value = newval
                    elif (field == "C"):
                        ws["G" + str(emprow)].value = newval
                wb.save(filename) # save file
                print("Changes applied!")
                while True:
                    curedit = input("Do you wish to edit further for the current employee (Y or N): ")
                    if (curedit not in yesorno):
                        print("Invalid value, please try again")
                        continue
                    break
                continue
    
        while True: # if not editing current emp anymore
                edit = input("Do you wish to exit to home screen(HS) or edit another employee's record(DE): ") # change edit
                if (edit not in edits):
                    print("Invalid value, please try again")
                    continue
                break           
    
    wb.close() # close file
    return

def change_pw():
    while True:
        pw = input("Please enter new password (enter -1 to quit): ")
        pw = pw.strip()
        if pw == "-1":
            return ""
        elif pw == "":
            print("Empty password not allowed, please try again")
            continue
        confirmpw = input("Please confirm new password (enter -1 to quit): ")
        confirmpw = confirmpw.strip()
        if confirmpw == "-1":
            return ""
        if confirmpw == pw:
            print("Password changed to:", confirmpw)
            break
        print("Passwords do not match, please try again")
    return confirmpw

# set password for excel file as same one you entered, password from welcome function
def setPW(pw):
    xlApp = w3c.Dispatch("Excel.Application") #using win32com again to set password
    wb = xlApp.Workbooks.Open(filename, False, False, None, "") #open without password
    xlApp.DisplayAlerts = False
    wb.SaveAs(filename, None, pw, "") # save with pw
    xlApp.Quit()
    return

####@ You update the codes here
def functionMenu():
    print(80*"-")
    print()
    print("""\
Functions available are:
    1. View Birthday
    2. Add Employee Record
    3. Remove Employee Record
    4. Edit Employee Record
    5. Change password
    x. Exit
""")
    return

def processCommand():
    ####@ You update the codes here
    pw = ""
    while True:
        functionMenu()
        cmdStr = input("Enter command character to proceed: ")
        cmdStr = cmdStr.strip()
        if cmdStr == "1":
            ####@ Do command 1
            print()
            print_bdays(get_bdays())
            print()
        elif cmdStr == "2":
            ####@ Do command 2
            print()
            add_record()
            print()
        elif cmdStr == "3":
            ####@ Do command 3
            print("-"*80)
            print()
            remove_record()
            print()
        elif cmdStr == "4":
            ####@ Do command 4
            print("-"*80)
            print()
            edit_record()
            print()
        elif cmdStr == "5":
#           ####@ Do command 5
            print("-"*80)
            print()
            pw = change_pw()
            print()
        elif cmdStr == "x":
            print()
            txt = "Thanks for using the HR Birthday Program. See you again!"
            print(txt.center(80, "-"))
            break
        else:
            ####@ Do something
            print(f"\n\nUnknown command '{cmdStr}' entered.  Please try again.\n")
    ###
    return pw

def main():
    access, pw = welcome() # return values are stored in access and pw, to check if access == 1 (allowed) and to save pw again later
    newpw = ""
    if access == 1: # access granted
        try:
            newpw = processCommand() #1235 for HR Melissa Tan, pw is lloyd
        except Exception as e:
            print(e)
            if newpw == "":
                setPW(pw) # if error, ensure password is saved again
            else:
                setPW(newpw)
            return
        finally:
            if newpw == "":
                setPW(pw)
            else:
                setPW(newpw) # save password before program ends
    return
    

if (__name__ == "__main__"):
    main()